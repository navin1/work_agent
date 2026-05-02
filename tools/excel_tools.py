"""Excel/DuckDB tools — ingestion and querying of mapping and master files."""
import json
import re
from core.json_utils import safe_json
import time
from datetime import datetime
from pathlib import Path

from langchain.tools import tool

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core import config, persistence
from core.audit import log_audit
from core.duckdb_manager import get_manager
from core.sql_formatter import extract_sql

# ── Validation export styling ─────────────────────────────────────────────────

_VX_HEADER_FILL     = PatternFill("solid", fgColor="4472C4")
_VX_HEADER_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_VX_DATA_FONT       = Font(name="Courier New", size=11)
_VX_STATUS_EMOJI: dict[str, str] = {
    "PASS":           "🟢 PASS",
    "FAIL":           "🔴 FAIL",
    "PARTIAL":        "🟡 PARTIAL",
    "NOT_APPLICABLE": "⚪ N/A",
    "NOT_EVALUATED":  "🔵 NOT EVALUATED",
    "ERROR":          "🟣 ERROR",
}
_VX_PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
_VX_FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
_VX_ALT_FILL   = PatternFill("solid", fgColor="F5F5F5")
_VX_CENTER     = Alignment(horizontal="center", vertical="center")
_VX_WRAP       = Alignment(horizontal="left",   vertical="top", wrap_text=True)


def _vx_safe_tab(name: str, used: set) -> str:
    name = re.sub(r'[\\/*?:\[\]]', "_", name)[:31]
    base, i = name, 2
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    return name


def _vx_header_row(ws, headers: list) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _VX_HEADER_FILL
        c.font = _VX_HEADER_FONT
        c.alignment = _VX_CENTER


def _vx_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_validation_excel(
    results: list,
    env_label: str,
    output_dir,
) -> Path:
    """Build Mapping_Results_<ENV>_<YYYYMMDD_HHMMSS>.xlsx from validation results.

    Args:
        results:    List of _do_validate_mapping result dicts.
        env_label:  Environment label used in filename (local / git / qa / prod …).
        output_dir: Directory to write the file into.

    Returns:
        Path to the written .xlsx file.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"Mapping_Results_{env_label}_{ts}.xlsx"

    wb       = openpyxl.Workbook()
    used     : set = set()

    # ── Summary tab ───────────────────────────────────────────────────────────
    ws_sum       = wb.active
    ws_sum.title = "Summary"
    used.add("Summary")

    _vx_header_row(ws_sum, ["File Name", "Pass", "Fail", "Partial", "Not Applicable", "Not Evaluated", "Total"])
    ws_sum.row_dimensions[1].height = 20

    for r, res in enumerate(results, 2):
        s    = res.get("summary", {})
        name = Path(res.get("mapping_file", "unknown")).stem
        for c, val in enumerate([
            name,
            s.get("pass", 0), s.get("fail", 0), s.get("partial", 0),
            s.get("not_applicable", 0), s.get("not_evaluated", 0), s.get("total", 0),
        ], 1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            cell.alignment = _VX_CENTER if c > 1 else Alignment(horizontal="left")
            if r % 2 == 0:
                cell.fill = _VX_ALT_FILL
        ws_sum.cell(row=r, column=2).fill = _VX_PASS_FILL
        ws_sum.cell(row=r, column=3).fill = _VX_FAIL_FILL

    _vx_col_widths(ws_sum, [35, 8, 8, 10, 16, 16, 8])
    ws_sum.freeze_panes = "A2"

    # ── Per-file tabs ─────────────────────────────────────────────────────────
    for res in results:
        stem     = Path(res.get("mapping_file", "unknown")).stem
        tab      = _vx_safe_tab(stem, used)
        used.add(tab)
        ws       = wb.create_sheet(title=tab)
        _vx_header_row(ws, ["Column", "Status", "Confidence", "Reason", "Evidence", "SQL File"])
        ws.row_dimensions[1].height = 20

        row = 2
        for group in res.get("bq_table_groups", []):
            for rule in group.get("rules", []):
                verdict       = (rule.get("verdict") or "").upper()
                status_label  = _VX_STATUS_EMOJI.get(verdict, verdict)
                afill         = _VX_ALT_FILL if row % 2 == 0 else None

                for col_name in (rule.get("target_columns") or [""]):
                    for ci, (val, aln) in enumerate([
                        (col_name,                       Alignment(horizontal="left")),
                        (status_label,                   _VX_CENTER),
                        (rule.get("confidence_tier",""), _VX_CENTER),
                        (rule.get("reason",         ""), _VX_WRAP),
                        (rule.get("evidence",       ""), _VX_WRAP),
                        (rule.get("sql_file",       ""), _VX_WRAP),
                    ], 1):
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.alignment = aln
                        cell.font = _VX_DATA_FONT
                        if afill:
                            cell.fill = afill
                    ws.row_dimensions[row].height = 30
                    row += 1

        _vx_col_widths(ws, [28, 16, 13, 65, 55, 40])
        ws.freeze_panes = "A2"

    wb.save(out_path)
    return out_path


def _dag_meta_for_path(file_path: str) -> dict:
    """Return {bq_table, dag_names, mapping_columns} for a file stem from excel_mapping.json.
    Always reads the live JSON, never trusts stale registry values."""
    stem = Path(file_path).stem
    excel_map = persistence.get_excel_mapping()
    return excel_map.get(stem) or excel_map.get(stem.lower()) or {}


def _safe_table_name(folder: str, stem: str) -> str:
    import re
    name = f"{folder}_{stem}".lower()
    name = re.sub(r"[^a-z0-9_]", "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def _cast_arrow_to_string(arrow) -> "pa.Table":
    """Cast every column in a PyArrow table to VARCHAR/string.

    Three-level fallback per column:
      1. pyarrow.compute.cast  — fast, handles numeric/temporal types
      2. stringify via to_pylist — handles binary, nested, and exotic types
      3. null array             — last resort so ingestion never aborts
    """
    import pyarrow as pa
    import pyarrow.compute as pc

    string_arrays = []
    for col in arrow.columns:
        try:
            string_arrays.append(pc.cast(col, pa.string(), safe=False))
        except Exception:
            try:
                string_arrays.append(pa.array(
                    [str(v) if v is not None else None for v in col.to_pylist()],
                    type=pa.string(),
                ))
            except Exception:
                string_arrays.append(pa.nulls(len(col), type=pa.string()))

    return pa.table(dict(zip(arrow.column_names, string_arrays)))


def _ingest_file(path: Path, folder: str) -> dict | None:
    """Ingest one Excel file into DuckDB. Returns registry entry or None on failure."""
    try:
        import pyarrow as pa
        stat = path.stat()
        stem = path.stem
        table_name = _safe_table_name(folder, stem)

        is_master = folder.lower() == "master"

        if is_master:
            try:
                import polars as pl
                try:
                    df_full = pl.read_excel(str(path), infer_schema_length=0)
                except TypeError:
                    df_full = pl.read_excel(str(path), read_options={"infer_schema_length": 0})
                try:
                    df_full = df_full.select(pl.all().cast(pl.String, strict=False))
                except Exception:
                    try:
                        df_full = df_full.select(pl.all().cast(pl.Utf8, strict=False))
                    except Exception:
                        pass
                row_count = len(df_full)
                if row_count > config.LARGE_FILE_ROW_THRESHOLD:
                    arrow = df_full.to_arrow()
                else:
                    import pandas as pd
                    df_pd = pd.read_excel(str(path), header=0, dtype=str)
                    arrow = pa.Table.from_pandas(df_pd, preserve_index=False)
            except ImportError:
                import pandas as pd
                df_pd = pd.read_excel(str(path), header=0, dtype=str)
                arrow = pa.Table.from_pandas(df_pd, preserve_index=False)

            bq_table = ""
            dag_names: list[str] = []
        else:
            # Both bq_table and dag_names come from excel_mapping.json keyed by file stem
            # Rows 1-3 = metadata, row 4 = headers, row 5+ = data
            import pandas as pd
            excel_map = persistence.get_excel_mapping()
            entry_meta = excel_map.get(stem) or excel_map.get(stem.lower()) or {}
            bq_table = entry_meta.get("bq_table", "")
            if isinstance(bq_table, list):
                bq_table = ", ".join(bq_table)
            dag_names = entry_meta.get("dag_names", [])

            try:
                import polars as pl
                try:
                    df_check = pl.read_excel(str(path), infer_schema_length=0, read_options={"skip_rows": 3})
                except TypeError:
                    df_check = pl.read_excel(str(path), read_options={"skip_rows": 3, "infer_schema_length": 0})
                if len(df_check) > config.LARGE_FILE_ROW_THRESHOLD:
                    try:
                        df_check = df_check.select(pl.all().cast(pl.String, strict=False))
                    except Exception:
                        try:
                            df_check = df_check.select(pl.all().cast(pl.Utf8, strict=False))
                        except Exception:
                            pass
                    arrow = df_check.to_arrow()
                    row_count = len(df_check)
                else:
                    raise ValueError("use pandas")
            except Exception:
                data_df = pd.read_excel(str(path), header=3, dtype=str)
                row_count = len(data_df)
                arrow = pa.Table.from_pandas(data_df, preserve_index=False)

        arrow = _cast_arrow_to_string(arrow)
        get_manager().register_table(table_name, arrow)

        entry = {
            "table_name": table_name,
            "bq_table": bq_table,
            "dag_names": dag_names,
            "source_folder": folder,
            "file_path": str(path),
            "columns": [field.name for field in arrow.schema],
            "row_count": len(arrow),
            "last_ingested": time.time(),
            "file_mtime": stat.st_mtime,
        }
        return entry
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("Failed to ingest %s: %s", path, exc)
        return None


def ingest_excel_files(folder_filter: str = None) -> dict:
    """Ingest all Excel files from data/mapping and data/master into DuckDB.
    Silently skips if no mapping or master directories exist."""
    data_root = Path(config.DATA_ROOT)
    mapping_root = data_root / "mapping"
    master_root = data_root / "master"

    registry = list(persistence.get_registry())
    registry_index = {e["file_path"]: e for e in registry}

    loaded = []
    skipped = []
    errors = []

    def process_folder(folder_path: Path, folder_name: str):
        if not folder_path.exists():
            return
        for xlsx in folder_path.rglob("*.xlsx"):
            if folder_filter and folder_filter.lower() not in str(xlsx).lower():
                continue
            existing = registry_index.get(str(xlsx))
            if existing and existing.get("file_mtime") == xlsx.stat().st_mtime:
                _ingest_file(xlsx, folder_name)
                skipped.append(xlsx.name)
                continue
            entry = _ingest_file(xlsx, folder_name)
            if entry:
                registry_index[str(xlsx)] = entry
                loaded.append(xlsx.name)
            else:
                errors.append(xlsx.name)

    if mapping_root.exists():
        for xlsx in mapping_root.rglob("*.xlsx"):
            if folder_filter and folder_filter.lower() not in str(xlsx).lower():
                continue
            folder_name = xlsx.parent.name if xlsx.parent != mapping_root else "mapping"
            existing = registry_index.get(str(xlsx))
            if existing and existing.get("file_mtime") == xlsx.stat().st_mtime:
                _ingest_file(xlsx, folder_name)
                meta = _dag_meta_for_path(str(xlsx))
                if meta:
                    bq_t = meta.get("bq_table", existing.get("bq_table", ""))
                    if isinstance(bq_t, list):
                        bq_t = ", ".join(bq_t)
                    existing["bq_table"] = bq_t
                    existing["dag_names"] = meta.get("dag_names", existing.get("dag_names", []))
                    registry_index[str(xlsx)] = existing
                skipped.append(xlsx.name)
                continue
            entry = _ingest_file(xlsx, folder_name)
            if entry:
                registry_index[str(xlsx)] = entry
                loaded.append(xlsx.name)
            else:
                errors.append(xlsx.name)

    process_folder(master_root, "master")

    updated_registry = list(registry_index.values())
    persistence.save_registry(updated_registry)

    return {"loaded": len(loaded), "skipped": len(skipped), "errors": errors, "files": loaded}


def _get_loaded_tables_internal() -> list[dict]:
    registry = persistence.get_registry()
    db = get_manager()
    active = set(db.list_tables())
    result = []
    for entry in registry:
        if entry["table_name"] in active:
            result.append(entry)
    return result


@tool
def query_excel_data(sql: str) -> str:
    """Execute SQL against DuckDB containing all ingested Excel mapping and master files.
    Table names follow pattern {folder}_{filename} e.g. rps800_reconciliation, master_products.
    Use list_loaded_tables first to confirm table names. Returns JSON with columns and rows.
    Returns empty result (not error) if no tables are loaded yet."""
    start = time.time()
    try:
        from core.sql_formatter import is_ddl_dml, format_sql
        if is_ddl_dml(sql):
            return json.dumps({"error": "DDL/DML not permitted. Only SELECT queries are allowed."})
        db = get_manager()
        if not db.list_tables():
            ingest_excel_files()
        if not db.list_tables():
            return json.dumps({
                "columns": [], "rows": [], "row_count": 0,
                "note": "No Excel tables loaded. Check that data/mapping/ has .xlsx files.",
            })
        df = db.execute(sql)
        duration = int((time.time() - start) * 1000)
        result = {
            "columns": list(df.columns),
            "rows": df.values.tolist(),
            "row_count": len(df),
            "formatted_sql": format_sql(sql),
        }
        log_audit("excel_tools", "duckdb", sql, row_count=len(df), duration_ms=duration)
        return safe_json(result)
    except Exception as exc:
        log_audit("excel_tools", "duckdb", sql, duration_ms=int((time.time() - start) * 1000))
        return json.dumps({"error": str(exc)})


@tool
def list_loaded_tables(folder_filter: str = None) -> str:
    """List all Excel files loaded into DuckDB.
    Returns JSON list with: table_name, source_folder, row_count,
    bq_table_reference, dag_names, last_ingested, file_path.
    Returns empty list (not error) if no Excel files are configured."""
    try:
        tables = _get_loaded_tables_internal()
        if not tables:
            # DuckDB is in-memory — re-ingest from disk transparently on each fresh session
            ingest_excel_files()
            tables = _get_loaded_tables_internal()
        if folder_filter:
            tables = [t for t in tables if folder_filter.lower() in t.get("source_folder", "").lower()]
        log_audit("excel_tools", "duckdb", "list_loaded_tables", row_count=len(tables))
        if not tables:
            return json.dumps({
                "tables": [],
                "count": 0,
                "note": "No Excel mapping files are loaded. Place .xlsx files in data/mapping/<folder>/",
            })
        return safe_json(tables)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


@tool
def get_table_schema(table_name: str) -> str:
    """Get column names and types for a DuckDB table.
    Call this before writing queries against unfamiliar tables.
    Returns JSON schema."""
    try:
        db = get_manager()
        if not db.list_tables():
            return json.dumps({"error": "No Excel tables loaded.", "table_name": table_name})
        schema = db.get_schema(table_name)
        log_audit("excel_tools", "duckdb", f"schema:{table_name}")
        return json.dumps({"table_name": table_name, "columns": schema})
    except Exception as exc:
        return json.dumps({"error": str(exc)})


@tool
def get_bq_table_for_mapping_file(mapping_file_name: str) -> str:
    """Return the BigQuery table(s) a mapping file maps to (sourced from excel_mapping.json).
    Returns BQ table reference string or 'not found'."""
    try:
        registry = persistence.get_registry()
        if not registry:
            return json.dumps({"bq_table": "not found", "note": "No Excel mapping files loaded."})
        name_lower = mapping_file_name.lower()
        for entry in registry:
            if name_lower in entry.get("file_path", "").lower() or name_lower in entry.get("table_name", "").lower():
                log_audit("excel_tools", "registry", f"bq_table_for:{mapping_file_name}")
                meta = _dag_meta_for_path(entry.get("file_path", ""))
                bq_table = meta.get("bq_table") or entry.get("bq_table") or "not found"
                if isinstance(bq_table, list):
                    bq_table = ", ".join(bq_table)
                return json.dumps({"bq_table": bq_table})
        return json.dumps({"bq_table": "not found"})
    except Exception as exc:
        return json.dumps({"error": str(exc)})


@tool
def get_dags_for_mapping_file(mapping_file_name: str) -> str:
    """Return DAG names associated with a mapping file (sourced from excel_mapping.json).
    Returns JSON list of DAG name strings."""
    try:
        registry = persistence.get_registry()
        if not registry:
            return json.dumps([])
        name_lower = mapping_file_name.lower()
        for entry in registry:
            if name_lower in entry.get("file_path", "").lower() or name_lower in entry.get("table_name", "").lower():
                log_audit("excel_tools", "registry", f"dags_for:{mapping_file_name}")
                meta = _dag_meta_for_path(entry.get("file_path", ""))
                dag_names = meta.get("dag_names") or entry.get("dag_names", [])
                return json.dumps(dag_names)
        return json.dumps([])
    except Exception as exc:
        return json.dumps({"error": str(exc)})


@tool
def reingest_excel_files(folder_filter: str = None) -> str:
    """Re-ingest Excel files from disk into DuckDB.
    Only re-ingests files modified since last ingest (mtime check).
    Optionally scoped to a specific folder. Returns ingest summary.
    Safe to call even if no Excel files exist."""
    try:
        result = ingest_excel_files(folder_filter=folder_filter)
        log_audit("excel_tools", "disk", "reingest", row_count=result["loaded"])
        return json.dumps(result)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


@tool
def trace_from_excel(mapping_file_name: str, composer_env: str = None) -> str:
    """Full end-to-end trace starting from an Excel mapping file.
    Given an Excel file name, returns: BigQuery table, DAG names, Airflow job (run) details,
    task list, rendered SQL for each task, and execution log summary.
    This is the primary tool for tracing data lineage from an Excel mapping sheet.

    mapping_file_name: partial name or table_name of the Excel mapping file.
    composer_env: optional Composer environment to query. If not provided, uses pinned workspace
                  or first available environment.
    Returns JSON with full trace: bq_table, dag_names, airflow_jobs, tasks, rendered_sqls."""
    start = time.time()
    try:
        registry = persistence.get_registry()
        if not registry:
            return json.dumps({
                "status": "no_excel",
                "note": "No Excel mapping files are loaded. Place .xlsx files in data/mapping/<folder>/",
            })

        # Find matching registry entry
        name_lower = mapping_file_name.lower()
        entry = None
        for e in registry:
            if (name_lower in e.get("file_path", "").lower() or
                    name_lower in e.get("table_name", "").lower() or
                    name_lower in Path(e.get("file_path", "")).stem.lower()):
                entry = e
                break

        if not entry:
            return json.dumps({
                "status": "not_found",
                "note": f"No mapping file matching '{mapping_file_name}' found in registry.",
                "available": [e.get("table_name") for e in registry],
            })

        meta = _dag_meta_for_path(entry.get("file_path", ""))
        bq_table = meta.get("bq_table") or entry.get("bq_table", "")
        if isinstance(bq_table, list):
            bq_table = "\n".join(bq_table)
        dag_names = meta.get("dag_names") or entry.get("dag_names", [])

        result = {
            "excel_file": Path(entry.get("file_path", "")).name,
            "table_name": entry.get("table_name"),
            "source_folder": entry.get("source_folder"),
            "bq_table": bq_table,
            "dag_names": dag_names,
            "composer_env": "",
            "airflow_jobs": [],
            "dag_details": [],
            "note": None,
        }

        if not dag_names:
            result["note"] = "No DAG names found for this mapping file in excel_mapping.json."
            return json.dumps(result)

        # Resolve composer env
        env = composer_env
        if not env:
            from core.workspace import get_pinned_workspace
            ws = get_pinned_workspace()
            env = ws.get("composer_env")
        if not env and config.COMPOSER_ENVS:
            env = next(iter(config.COMPOSER_ENVS))

        if not env or env not in config.COMPOSER_ENVS:
            result["note"] = (
                f"Composer environment not configured or not found. "
                f"Available: {list(config.COMPOSER_ENVS.keys()) or 'none'}. "
                "Pass composer_env parameter or configure COMPOSER_ENVS in .env"
            )
            return json.dumps(result)

        result["composer_env"] = env

        # For each DAG: get recent jobs + task graph + rendered SQL
        from tools.composer_tools import _get
        from core.sql_formatter import format_sql

        for dag_id in dag_names:
            dag_info = {
                "dag_id": dag_id,
                "recent_jobs": [],
                "tasks": [],
                "rendered_sqls": [],
                "error": None,
            }

            try:
                # Recent runs (jobs)
                runs_data = _get(env, f"/dags/{dag_id}/dagRuns",
                                 {"limit": 5, "order_by": "-start_date"})
                latest_run_id = None
                for r in runs_data.get("dag_runs", []):
                    s = r.get("start_date")
                    e = r.get("end_date")
                    duration = None
                    if s and e:
                        try:
                            from datetime import datetime
                            duration = (
                                datetime.fromisoformat(e.replace("Z", "+00:00")) -
                                datetime.fromisoformat(s.replace("Z", "+00:00"))
                            ).total_seconds()
                        except Exception:
                            pass
                    dag_info["recent_jobs"].append({
                        "run_id": r.get("dag_run_id"),
                        "state": r.get("state"),
                        "start_time": s,
                        "end_time": e,
                        "duration_seconds": duration,
                    })
                    if latest_run_id is None:
                        latest_run_id = r.get("dag_run_id")

                # Task list with states from latest run
                tasks_data = _get(env, f"/dags/{dag_id}/tasks")
                task_defs = {t["task_id"]: t for t in tasks_data.get("tasks", [])}

                task_states = {}
                if latest_run_id:
                    try:
                        instances = _get(env, f"/dags/{dag_id}/dagRuns/{latest_run_id}/taskInstances")
                        for inst in instances.get("task_instances", []):
                            task_states[inst["task_id"]] = {
                                "state": inst.get("state"),
                                "duration_seconds": inst.get("duration"),
                            }
                    except Exception:
                        pass

                for tid, tdef in task_defs.items():
                    state_info = task_states.get(tid, {})
                    dag_info["tasks"].append({
                        "task_id": tid,
                        "operator": tdef.get("class_ref", {}).get("class_name", ""),
                        "depends_on": tdef.get("downstream_task_ids", []),
                        "state": state_info.get("state"),
                        "duration_seconds": state_info.get("duration_seconds"),
                    })

                # Rendered SQL for each task — use same approach as get_task_sql
                from tools.composer_tools import (  # noqa: PLC0415
                    _extract_rendered_sql, _best_sql, _enc,
                    _get_sql_file_path, _rendered_was_truncated, _fetch_sql_file,
                )
                success_runs_list = []
                try:
                    success_runs = _get(env, f"/dags/{dag_id}/dagRuns",
                                        {"limit": 10, "order_by": "-execution_date", "state": "success"})
                    success_runs_list = success_runs.get("dag_runs", [])
                except Exception:
                    pass

                for tid in task_defs:
                    raw_sql = None
                    rendered_sql = None
                    rendered_truncated = False

                    # Raw SQL from task definition — use _get_sql_file_path to handle
                    # BigQueryInsertJobOperator's nested configuration.query.query path
                    try:
                        task_data = _get(env, f"/dags/{dag_id}/tasks/{tid}")
                        sql_file_path = _get_sql_file_path(task_data)
                        if sql_file_path:
                            raw_sql = _fetch_sql_file(sql_file_path)
                        if not raw_sql:
                            raw_sql = extract_sql(task_data)
                    except Exception:
                        pass

                    # Rendered SQL — walk recent runs until this task instance is found
                    for dag_run in success_runs_list:
                        run_id = dag_run["dag_run_id"]
                        try:
                            ti_detail = _get(env,
                                             f"/dags/{_enc(dag_id)}/dagRuns/{_enc(run_id)}/taskInstances/{_enc(tid)}")
                            rendered_sql = _extract_rendered_sql(ti_detail)
                            rendered_truncated = _rendered_was_truncated(ti_detail)
                            if rendered_sql:
                                break
                        except Exception:
                            continue

                    best = _best_sql(raw_sql, rendered_sql, rendered_truncated)
                    if best:
                        dag_info["rendered_sqls"].append({
                            "task_id": tid,
                            "rendered_sql": format_sql(best).replace("\xa0", " ").replace("\\xa0", " "),
                        })

            except Exception as e:
                dag_info["error"] = str(e)

            result["dag_details"].append(dag_info)

        log_audit("excel_tools", "trace", f"trace_from_excel:{mapping_file_name}",
                  duration_ms=int((time.time()-start)*1000))
        return safe_json(result)
    except Exception as exc:
        return json.dumps({"error": str(exc)})
