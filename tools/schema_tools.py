"""Schema introspection and MySQL→BigQuery schema audit tools."""
import json
from core.json_utils import safe_json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from langchain.tools import tool

from core import config
from core.audit import log_audit


# ── BQ schema introspection helpers ──────────────────────────────────────────

def _bq_field_to_dict(field, prefix: str = "") -> dict:
    path = f"{field.name}" if not prefix else f"{prefix}.{field.name}"
    result = {
        "name": field.name,
        "path": path,
        "field_type": field.field_type,
        "mode": field.mode,
        "description": field.description or "",
        "fields": [],
    }
    if field.field_type in ("RECORD", "STRUCT") and field.fields:
        for subfield in field.fields:
            result["fields"].append(_bq_field_to_dict(subfield, prefix=path))
    return result


def _flatten_fields(field_dict: dict, leaf_list: list) -> None:
    if not field_dict.get("fields"):
        leaf_list.append({
            "path": field_dict["path"],
            "field_type": field_dict["field_type"],
            "mode": field_dict["mode"],
        })
    else:
        for sub in field_dict["fields"]:
            _flatten_fields(sub, leaf_list)


@tool
def introspect_bq_schema(project_id: str, dataset_id: str, table_id: str) -> str:
    """Recursively fetch full BigQuery table schema to ALL nesting levels.
    NEVER truncates RECORD/STRUCT/REPEATED fields. Recurses until no further nesting exists.
    Returns JSON with nested tree AND flat list of all leaf fields with full dot-path."""
    start = time.time()
    try:
        from tools.bigquery_tools import _validate_project, _get_client
        err = _validate_project(project_id)
        if err:
            return json.dumps({"error": err})
        client = _get_client()
        table_ref = f"{project_id}.{dataset_id}.{table_id}"
        table = client.get_table(table_ref)

        tree = [_bq_field_to_dict(f) for f in table.schema]
        leaves: list[dict] = []
        for node in tree:
            _flatten_fields(node, leaves)

        result = {
            "table": table_ref,
            "row_count": table.num_rows,
            "size_bytes": table.num_bytes,
            "schema_tree": tree,
            "flat_fields": leaves,
            "field_count": len(leaves),
        }
        log_audit("schema_tools", table_ref, "introspect_bq_schema",
                  duration_ms=int((time.time()-start)*1000))
        return safe_json(result)
    except Exception as exc:
        return json.dumps({"error": str(exc)})


# ── Schema audit — MySQL metadata → BigQuery reconciliation ──────────────────

# MySQL base type → expected BigQuery type
_MYSQL_TO_BQ: dict[str, str] = {
    "int": "INT64", "bigint": "INT64", "smallint": "INT64",
    "tinyint": "INT64", "mediumint": "INT64",
    "float": "FLOAT64", "double": "FLOAT64",
    "decimal": "NUMERIC",
    "varchar": "STRING", "char": "STRING", "text": "STRING",
    "longtext": "STRING", "mediumtext": "STRING", "tinytext": "STRING",
    "datetime": "DATETIME", "timestamp": "TIMESTAMP",
    "date": "DATE", "time": "TIME",
    "boolean": "BOOL", "bool": "BOOL", "bit": "BOOL",
    "json": "JSON", "enum": "STRING",
    "blob": "BYTES", "longblob": "BYTES", "mediumblob": "BYTES",
}

# BigQuery Standard SQL aliases → canonical form
_BQ_ALIASES: dict[str, str] = {
    "INTEGER": "INT64", "INT": "INT64", "SMALLINT": "INT64",
    "BIGINT": "INT64", "BYTEINT": "INT64", "TINYINT": "INT64",
    "FLOAT": "FLOAT64", "STRING": "JSON",
    "DECIMAL": "NUMERIC", "BIGDECIMAL": "NUMERIC", "BIGNUMERIC": "NUMERIC",
    "BOOLEAN": "BOOL", "TIMESTAMP": "DATETIME",
}

_AUDIT_COLUMNS = [
    "Column Name", "MySQL #", "BQ #",
    "MySQL Type", "Expected BQ Type", "Actual BQ Type",
    "BQ Description", "Status",
]


def _mysql_to_bq(mysql_type: str) -> str:
    return _MYSQL_TO_BQ.get(mysql_type.strip().lower(), mysql_type.strip().upper())


def _canonical_bq(bq_type: str) -> str:
    return _BQ_ALIASES.get(bq_type.strip().upper(), bq_type.strip().upper())


def _audit_status(m: Optional[dict], b: Optional[dict], expected: str, actual: str) -> str:
    if m is None:
        return "🟠 BQ Only"
    if b is None:
        return "🔵 MySQL Only"
    if actual and expected and _canonical_bq(actual) != _canonical_bq(expected):
        return "🟡 Type Mismatch"
    return "🟢 Match"


def _safe_sheet_name(name: str, used: set) -> str:
    cleaned = re.sub(r"[\\\/\?\*\[\]:]", "_", name)[:31]
    if cleaned not in used:
        return cleaned
    for i in range(2, 1000):
        candidate = f"{cleaned[:31 - len(str(i)) - 1]}_{i}"
        if candidate not in used:
            return candidate
    return cleaned


def _fetch_mysql_metadata(client) -> "pd.DataFrame":
    import pandas as pd
    sql = f"""
    SELECT
        h.table_name, h.eda_dataset_name, h.eda_view_name, h.deployed_to_prod,
        d.column_name, CAST(d.ordinal_position AS INT64) AS ordinal_position, d.data_type
    FROM `{config.SCHEMA_HEADER_VIEW}` h
    JOIN `{config.SCHEMA_DETAIL_VIEW}` d ON d.table_name = h.table_name
    WHERE h.is_streamed = 1
    ORDER BY h.table_name, d.ordinal_position
    """
    df = client.query(sql).to_dataframe()
    return df


def _fetch_bq_schema(client, project: str, dataset: str, view: str) -> list[dict]:
    full_ref = f"{project}.{dataset}.{view}"
    try:
        table_ref = client.get_table(full_ref)
        return [
            {
                "column_name":      field.name,
                "ordinal_position": idx,
                "data_type":        field.field_type,
                "mode":             field.mode,
                "description":      field.description or "",
            }
            for idx, field in enumerate(table_ref.schema, 1)
        ]
    except Exception:
        return []


def _reconcile(mysql_rows: list[dict], bq_rows: list[dict]) -> list[dict]:
    mysql_map = {r["column_name"]: r for r in mysql_rows}
    bq_map    = {r["column_name"]: r for r in bq_rows}
    mysql_ordered = sorted(mysql_map, key=lambda c: mysql_map[c]["ordinal_position"])
    bq_only = sorted(
        (c for c in bq_map if c not in mysql_map),
        key=lambda c: bq_map[c]["ordinal_position"],
    )
    rows = []
    for col in mysql_ordered + bq_only:
        m = mysql_map.get(col)
        b = bq_map.get(col)
        mysql_type  = m["data_type"].strip() if m else ""
        actual_bq   = b["data_type"].strip() if b else ""
        expected_bq = _mysql_to_bq(mysql_type) if mysql_type else ""
        rows.append({
            "Column Name":      col,
            "MySQL #":          int(m["ordinal_position"]) if m else "",
            "BQ #":             int(b["ordinal_position"]) if b else "",
            "MySQL Type":       mysql_type,
            "Expected BQ Type": expected_bq,
            "Actual BQ Type":   actual_bq,
            "BQ Description":   b["description"] if b else "",
            "Status":           _audit_status(m, b, expected_bq, actual_bq),
        })
    return rows


def _apply_audit_sheet_format(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    status_fills = {
        "🟢": PatternFill("solid", fgColor="C6EFCE"),
        "🟡": PatternFill("solid", fgColor="FFEB9C"),
        "🟠": PatternFill("solid", fgColor="FF7518"),
        "🔵": PatternFill("solid", fgColor="BDD7EE"),
    }
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    data_font   = Font(name="Courier New", size=11)
    thin = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_col = None
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value == "Status":
            status_col = cell.column

    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            length = len(str(cell.value)) if cell.value is not None else 0
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), length)
            if cell.row > 1:
                cell.font = data_font
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center")
                if status_col and cell.column == status_col:
                    emoji = str(cell.value or "")[:2]
                    fill = status_fills.get(emoji)
                    if fill:
                        cell.fill = fill

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 60)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18


def _write_audit_excel(
    meta_df,
    tables: list[dict],
    bq_project: str,
    file_suffix: str,
    client,
    output_dir: str,
    timestamp: str,
) -> dict:
    import pandas as pd
    from openpyxl import load_workbook

    output_file = str(Path(output_dir) / f"schema_audit_{timestamp}_{file_suffix}.xlsx")
    all_rows: list[dict] = []
    table_dfs: dict[str, "pd.DataFrame"] = {}

    for tbl in tables:
        name    = tbl["table_name"]
        dataset = tbl["eda_dataset_name"]
        view    = tbl["eda_view_name"]
        mysql_rows = (
            meta_df[meta_df["table_name"] == name][
                ["column_name", "ordinal_position", "data_type"]
            ].to_dict("records")
        )
        bq_rows = _fetch_bq_schema(client, bq_project, dataset, view)
        rows = _reconcile(mysql_rows, bq_rows)
        table_dfs[name] = pd.DataFrame(rows, columns=_AUDIT_COLUMNS)
        for row in rows:
            all_rows.append({"Table Name": name, **row})

    if not all_rows:
        return {"error": f"No columns reconciled for {file_suffix}"}

    summary_rows = [r for r in all_rows if not r["Status"].startswith("🟢")]
    summary_df = pd.DataFrame(
        summary_rows if summary_rows else [{}],
        columns=["Table Name"] + _AUDIT_COLUMNS,
    )
    used_names: set = {"Summary"}
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        for tbl_name, df in table_dfs.items():
            sheet = _safe_sheet_name(tbl_name, used_names)
            used_names.add(sheet)
            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        _apply_audit_sheet_format(wb[sheet_name])
    wb.save(output_file)

    total      = len(all_rows)
    matches    = sum(1 for r in all_rows if r["Status"].startswith("🟢"))
    mismatches = sum(1 for r in all_rows if r["Status"].startswith("🟡"))
    bq_only    = sum(1 for r in all_rows if r["Status"].startswith("🟠"))
    mysql_only = sum(1 for r in all_rows if r["Status"].startswith("🔵"))
    return {
        "output_file":    output_file,
        "tables":         len(table_dfs),
        "total_columns":  total,
        "match":          matches,
        "type_mismatch":  mismatches,
        "bq_only":        bq_only,
        "mysql_only":     mysql_only,
    }


def _write_ddl_json(
    meta_df,
    tables: list[dict],
    file_suffix: str,
    output_dir: str,
    timestamp: str,
) -> str:
    output_file = str(Path(output_dir) / f"schema_ddl_{timestamp}_{file_suffix}.json")
    ddl: list[dict] = []
    for tbl in tables:
        name = tbl["table_name"]
        cols = (
            meta_df[meta_df["table_name"] == name]
            .sort_values("ordinal_position")[["column_name", "data_type"]]
            .to_dict("records")
        )
        ddl.append({
            "Table": name,
            "Schema": [
                {"name": c["column_name"], "type": _mysql_to_bq(c["data_type"]), "mode": "NULLABLE"}
                for c in cols
            ],
        })
    with open(output_file, "w", encoding="utf-8") as fh:
        json.dump(ddl, fh, indent=2)
    return output_file


def _write_audit_json(
    meta_df,
    tables: list[dict],
    bq_project: str,
    file_suffix: str,
    client,
    output_dir: str,
    timestamp: str,
) -> str:
    output_file = str(Path(output_dir) / f"schema_audit_{timestamp}_{file_suffix}.json")
    audit: list[dict] = []

    for tbl in tables:
        name    = tbl["table_name"]
        dataset = tbl["eda_dataset_name"]
        view    = tbl["eda_view_name"]

        mysql_rows = (
            meta_df[meta_df["table_name"] == name][
                ["column_name", "ordinal_position", "data_type"]
            ].to_dict("records")
        )
        bq_rows = _fetch_bq_schema(client, bq_project, dataset, view)
        rows    = _reconcile(mysql_rows, bq_rows)

        table_exists = len(bq_rows) > 0
        added: list[dict]      = []
        removed: list[dict]    = []
        mismatches: list[dict] = []

        for row in rows:
            status = row["Status"]
            col    = row["Column Name"]
            if status.startswith("🔵"):  # MySQL Only → added in MySQL, missing from BQ
                added.append({
                    "name": col,
                    "type": row["Expected BQ Type"],
                    "mode": "NULLABLE",
                })
            elif status.startswith("🟠"):  # BQ Only → removed from MySQL
                bq_info = next((b for b in bq_rows if b["column_name"] == col), {})
                removed.append({
                    "name": col,
                    "type": row["Actual BQ Type"],
                    "mode": bq_info.get("mode", "NULLABLE"),
                })
            elif status.startswith("🟡"):  # Type Mismatch
                mismatches.append({
                    "name":       col,
                    "BQ_Type":    row["Actual BQ Type"],
                    "MySQL_Type": row["Expected BQ Type"],
                })

        event_type = (
            (["added_columns"]      if added      else []) +
            (["removed_columns"]    if removed    else []) +
            (["datatype_mismatches"] if mismatches else [])
        )

        entry: dict = {
            "table_name":   name,
            "table_exists": str(table_exists).lower(),
            "event_type":   event_type,
        }
        if added:
            entry["added_columns"] = added
        if removed:
            entry["removed_columns"] = removed
        if mismatches:
            entry["datatype_mismatches"] = mismatches

        audit.append(entry)

    with open(output_file, "w", encoding="utf-8") as fh:
        json.dump(audit, fh, indent=2)
    return output_file


@tool
def run_schema_audit() -> str:
    """Run full MySQL → BigQuery schema reconciliation audit.

    Reads MySQL column metadata from BigQuery metadata views (SCHEMA_HEADER_VIEW,
    SCHEMA_DETAIL_VIEW), fetches actual BQ schemas, and reconciles them per column.

    Splits tables into prod (deployed_to_prod=1) and UAT batches.
    Generates for each batch:
      - schema_audit_<timestamp>_prd/uat.xlsx  — colour-coded Excel with Summary sheet
        and one sheet per table. Status: 🟢 Match · 🟡 Type Mismatch · 🟠 BQ Only · 🔵 MySQL Only
      - schema_ddl_<timestamp>_prd/uat.json    — BigQuery DDL JSON for all tables

    All files are saved to SCHEMA_AUDIT_OUTPUT_DIR (defaults to exports/).

    Required .env vars: SCHEMA_METADATA_PROJECT, SCHEMA_HEADER_VIEW, SCHEMA_DETAIL_VIEW.
    Optional: SCHEMA_BQ_PROJECT_PROD, SCHEMA_BQ_PROJECT_UAT.

    Returns JSON summary with file paths and column-level stats per batch."""
    start = time.time()
    try:
        if not config.SCHEMA_METADATA_PROJECT:
            return json.dumps({"error": "SCHEMA_METADATA_PROJECT is not set in .env"})
        if not config.SCHEMA_HEADER_VIEW or not config.SCHEMA_DETAIL_VIEW:
            return json.dumps({"error": "SCHEMA_HEADER_VIEW and SCHEMA_DETAIL_VIEW must be set in .env"})

        from google.cloud import bigquery
        from core.auth import get_credentials
        creds, _ = get_credentials()
        client = bigquery.Client(project=config.SCHEMA_METADATA_PROJECT, credentials=creds)

        meta_df = _fetch_mysql_metadata(client)
        if meta_df.empty:
            return json.dumps({"error": "No streamed tables found in SCHEMA_HEADER_VIEW"})

        tbl_cols    = ["table_name", "eda_dataset_name", "eda_view_name", "deployed_to_prod"]
        all_tables  = meta_df[tbl_cols].drop_duplicates().to_dict("records")
        prod_tables = [t for t in all_tables if t.get("deployed_to_prod") == 1]
        uat_tables  = [t for t in all_tables if t.get("deployed_to_prod") != 1]

        output_dir = config.SCHEMA_AUDIT_OUTPUT_DIR
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")

        results: dict = {
            "tables_found": len(all_tables),
            "prod_tables":  len(prod_tables),
            "uat_tables":   len(uat_tables),
        }

        if prod_tables and config.SCHEMA_BQ_PROJECT_PROD:
            results["prod"] = _write_audit_excel(
                meta_df, prod_tables, config.SCHEMA_BQ_PROJECT_PROD,
                "prd", client, output_dir, timestamp,
            )
            results["prod"]["ddl_json"] = _write_ddl_json(
                meta_df, prod_tables, "prd", output_dir, timestamp,
            )
            results["prod"]["audit_json"] = _write_audit_json(
                meta_df, prod_tables, config.SCHEMA_BQ_PROJECT_PROD,
                "prd", client, output_dir, timestamp,
            )
        elif prod_tables:
            results["prod_skipped"] = "SCHEMA_BQ_PROJECT_PROD not set"

        if uat_tables:
            uat_project = config.SCHEMA_BQ_PROJECT_UAT or config.SCHEMA_METADATA_PROJECT
            results["uat"] = _write_audit_excel(
                meta_df, uat_tables, uat_project,
                "uat", client, output_dir, timestamp,
            )
            results["uat"]["ddl_json"] = _write_ddl_json(
                meta_df, uat_tables, "uat", output_dir, timestamp,
            )
            results["uat"]["audit_json"] = _write_audit_json(
                meta_df, uat_tables, uat_project,
                "uat", client, output_dir, timestamp,
            )

        log_audit("schema_tools", config.SCHEMA_METADATA_PROJECT, "run_schema_audit",
                  row_count=len(all_tables),
                  duration_ms=int((time.time() - start) * 1000))
        return safe_json(results)
    except Exception as exc:
        return json.dumps({"error": str(exc)})
