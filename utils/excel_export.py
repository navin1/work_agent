"""Export mapping validation results to a styled Excel workbook."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Styling constants ─────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BOLD_FONT   = Font(bold=True)

_STATUS_FILL: dict[str, PatternFill] = {
    "PASS":            PatternFill("solid", fgColor="C6EFCE"),
    "FAIL":            PatternFill("solid", fgColor="FFC7CE"),
    "PARTIAL":         PatternFill("solid", fgColor="FFEB9C"),
    "NOT_APPLICABLE":  PatternFill("solid", fgColor="D9D9D9"),
    "NOT_EVALUATED":   PatternFill("solid", fgColor="BDD7EE"),
    "ERROR":           PatternFill("solid", fgColor="E2CFFF"),
}

_SUMMARY_PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
_SUMMARY_FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
_ALT_ROW_FILL       = PatternFill("solid", fgColor="F5F5F5")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_tab_name(name: str, used: set[str]) -> str:
    """Return an Excel-legal tab name (≤31 chars, no []?*/\\:, unique)."""
    name = re.sub(r'[\\/*?:\[\]]', "_", name)[:31]
    base, i = name, 2
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    return name


def _write_header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Public API ────────────────────────────────────────────────────────────────

def export_validation_excel(
    results: list[dict],
    env_label: str,
    output_dir: "Path | str",
) -> Path:
    """Build Mapping_Results_<ENV>_<YYYYMMDD_HHMMSS>.xlsx.

    Args:
        results:    List of _do_validate_mapping result dicts.
        env_label:  Short label for the environment (local / git / prod / qa …).
        output_dir: Directory to write the file.

    Returns:
        Path to the written file.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"Mapping_Results_{env_label}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    used_tabs: set[str] = set()

    # ── Tab 1: Summary ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    used_tabs.add("Summary")

    sum_headers = ["File Name", "Pass", "Fail", "Partial", "Not Applicable", "Not Evaluated", "Total"]
    _write_header_row(ws_sum, sum_headers)
    ws_sum.row_dimensions[1].height = 20

    for r_idx, res in enumerate(results, 2):
        s    = res.get("summary", {})
        name = Path(res.get("mapping_file", "unknown")).stem
        row_data = [
            name,
            s.get("pass",           0),
            s.get("fail",           0),
            s.get("partial",        0),
            s.get("not_applicable", 0),
            s.get("not_evaluated",  0),
            s.get("total",          0),
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = _CENTER if c_idx > 1 else Alignment(horizontal="left")
            if r_idx % 2 == 0:
                cell.fill = _ALT_ROW_FILL

        # Highlight pass/fail counts
        ws_sum.cell(row=r_idx, column=2).fill = _SUMMARY_PASS_FILL
        ws_sum.cell(row=r_idx, column=3).fill = _SUMMARY_FAIL_FILL

    _set_col_widths(ws_sum, [35, 8, 8, 10, 16, 16, 8])
    ws_sum.freeze_panes = "A2"

    # ── Tabs 2…N: one per validated file ──────────────────────────────────────
    detail_headers = ["Column", "Status", "Confidence", "Reason", "Evidence"]

    for res in results:
        file_stem = Path(res.get("mapping_file", "unknown")).stem
        tab_name  = _safe_tab_name(file_stem, used_tabs)
        used_tabs.add(tab_name)

        ws = wb.create_sheet(title=tab_name)
        _write_header_row(ws, detail_headers)
        ws.row_dimensions[1].height = 20

        row_idx = 2
        for group in res.get("bq_table_groups", []):
            for rule in group.get("rules", []):
                verdict    = (rule.get("verdict") or "").upper()
                status_fill = _STATUS_FILL.get(verdict)
                targets    = rule.get("target_columns") or [""]

                for col_name in targets:
                    # Alternate row shading (behind status fill for col 2)
                    base_fill = _ALT_ROW_FILL if row_idx % 2 == 0 else None

                    c1 = ws.cell(row=row_idx, column=1, value=col_name)
                    c1.alignment = Alignment(horizontal="left")
                    if base_fill:
                        c1.fill = base_fill

                    c2 = ws.cell(row=row_idx, column=2, value=verdict)
                    c2.alignment = _CENTER
                    if status_fill:
                        c2.fill = status_fill

                    c3 = ws.cell(row=row_idx, column=3, value=rule.get("confidence_tier", ""))
                    c3.alignment = _CENTER
                    if base_fill:
                        c3.fill = base_fill

                    c4 = ws.cell(row=row_idx, column=4, value=rule.get("reason", ""))
                    c4.alignment = _WRAP
                    if base_fill:
                        c4.fill = base_fill

                    c5 = ws.cell(row=row_idx, column=5, value=rule.get("evidence", ""))
                    c5.alignment = _WRAP
                    if base_fill:
                        c5.fill = base_fill

                    ws.row_dimensions[row_idx].height = 30
                    row_idx += 1

        _set_col_widths(ws, [28, 16, 13, 65, 55])
        ws.freeze_panes = "A2"

    wb.save(out_path)
    return out_path
